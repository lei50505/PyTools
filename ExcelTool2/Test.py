#! /usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import *
import openpyxl
import traceback
from openpyxl.styles import colors
from openpyxl.styles import Color,Font,Side,PatternFill,Border
from openpyxl.cell.read_only import EmptyCell,ReadOnlyCell
from openpyxl.workbook.workbook import Workbook
import time

class SrcCell():
    def __init__(self,val):
        self.val=val
        pass
    def getVal(self):
        return self.val

class SrcRow():
    def __init__(self):
        pass
    def getNumCell(self):
        pass
    def getVal(self):
        pass
    def EqualVal(self):
        pass

class SrcSheet():
    def __init__(self,sheet):
        self.rows=[]
        maxRow = sheet.max_row
        for index in range(1,maxRow+1):
            row = SrcRow()
            self.rows.append(SrcRow)
        self.rows = rows
        pass
    def getRow(self):
        pass
    def countNumCell(self):
        pass
    def setRowDone(self):
        pass
    def sort(self):
        pass
    def getDiffRows(self):
        pass
    def getFreeRows(self):
        pass
    def getDupliVals(self):
        pass
    def getRowsByVal(self):
        pass
    def getRowsInTotal(self):
        pass
    

class SrcBook():
    def __init__(self,path):
        if not os.path.isfile(path):
            raise MyError("源文件 %s 不存在" % path)
        book = openpyxl.load_workbook(path,read_only=True, \
                                    keep_vba=False, data_only=True, \
                                    guess_types=False, keep_links=False)
        
        self.sheet1 =SrcSheet(book["Sheet1"])
        self.sheet2 =SrcSheet(book["Sheet2"])
        pass
    def getSheetByName(self,sheetName):
        pass
    def addSheet(self):
        pass
    def findEqual(self):
        pass
    def compSheet(self):
        pass

class TarCell():
    def __init__(self,color):
        self.color = color
        pass
    def setColor(self,color):
        self.color=color
        pass

class TarRow():
    def __init__(self,cells):
        self.cells=cells
        pass
    def setColor(self,color):
        for cell in cells:
            cell.setColor(color)
        pass

class TarSheet():
    def __init__(self,rows):
        self.rows=rows
        pass
    def addRow(self,row):
        self.rows.append(row)
        row.setColor(color)
        pass

class TarBook():
    def __init__(self):

        pass
    def getSheet(self):
        pass
    def write(self,path):
        pass

def run():
    srcBook = SrcBook("in.xlsx")
    srcSheet1 = srcBook.getSheet()
    srcSheet2 = srcBook.getSheet()

    tarBook = TarBook()
    tarSheet = tarBook.getSheet()

    for row1 in sheet1.getDiffRows():
        for row2 in sheet2.getDiffRows():
            if row1.equalVal(row2):
                tarSheet.addRow(row1,color1)
                tarSheet.addRow(row2,color2)
                row1.done()
                row2.done()
    for val in sheet1.getDupliVals():

        if sheet1.countNumCell(val)==sheet2.countNumCell(val):
            for row1 in sheet1.getRowsByVal(row1Val):
                tarSheet.addRow(row1,color1)
                row1.done()
            for row2 in sheet1.getRowsByVal(row1Val):
                tarSheet.addRow(row2,color2)
                row2.done()

    for row1 in sheet1.getDiffRows():
        rows = sheet2.getRowsInTotal(row1.getVal())
        if len(rows) == 0:
            continue
        tarSheet.addRow(row1,color1)
        row1.done()
        for row in rows:
            tarSheet.addRow(row2,color2)
            row2.done()

    for row2 in sheet2.getDiffRows():
        rows = sheet1.getRowsInTotal(row2.getVal())
        if len(rows) == 0:
            continue
        tarSheet.addRow(row2,color2)
        row2.done()
        for row in rows:
            tarSheet.addRow(row1,color1)
            row1.done()
                

    for row1 in sheet1.getFreeRows():
        tarSheet.addRow(row1,color1)
        row1.done()

    for row2 in sheet2.getFreeRows():
        tarSheet.addRow(row2,color2)
        row2.done()


    tarBook.write(tarPath)
    

class ExcelStat():
    def __init__(self,index,val):
        self.index=index
        self.val=val
        self.done=False

class MyError(Exception):
    pass

class ExcelService(object):
    def __init__(self,srcPath="in.xlsx",tarPath="out.xlsx"):
        if not os.path.isfile(srcPath):
            raise MyError("源文件 %s 不存在" % srcPath)
        if os.path.isfile(tarPath):
            raise MyError("请删除目标文件 %s" % tarPath)
        self.srcPath = srcPath
        self.tarPath = tarPath

        self.srcBook = openpyxl.load_workbook(srcPath,read_only=True, \
                                    keep_vba=False, data_only=True, \
                                    guess_types=False, keep_links=False)
        self.tarBook = openpyxl.Workbook(write_only=True)
        

    def saveTarBook(self):
        self.tarBook.save(self.tarPath)

    def close(self):
        self.srcBook.close()
        self.tarBook.close()

        self.srcBook = None
        self.tarBook = None

        
        

def main():
    service = ExcelService()
    service.close()

def getDiffRows(sheet,col):
    rowValid = {}
    diffCells = []
    ret = []
    for row in range(1,sheet.max_row+1):
        rowValid[row]=True
        
    for row in range(1,sheet.max_row+1):
        cell = getCell(sheet,row,col)
        if type(cell) == EmptyCell:
            continue
        findFlag = False
        equalRow = None
        for diffCell in diffCells:
            if diffCell.value == cell.value:
                findFlag= True
                equalRow = diffCell.row
                break

        if findFlag ==False:
            diffCells.append(cell)
        else:
            rowValid[row]=False
            rowValid[equalRow]=False
    for row in range(1,sheet.max_row+1):
        if rowValid[row] == True:
            ret.append(row)
    return ret
            
    



def isSheetNamesInBook(book,*sheetNames):
    if len(sheetNames)==0:
        return False
    bookSheetNames = book.get_sheet_names()
    for sheetName in sheetNames:
        findFlag = False
        for bookSheetName in bookSheetNames:
            if sheetName == bookSheetName:
                findFlag=True
                break
        if findFlag == False:
            return False
    return True
        
    
def isNum(val):
    if val == None:
        return False
    try:
        int(val)
    except:
        try:
            float(val)
        except:
            return False
    return True

def getCellValIfNum(cell):
    
    if not isinstance(cell,ReadOnlyCell):
        return None

    val = cell.value

    if isinstance(val,float):
        return val

    if isinstance(val,int):
        return float(val)

    if isinstance(val,str):
        try:
            return float(val)
            
        except:

            return None
    
    return None


def getNumColIndex(sheet):

    maxCol = sheet.max_column
    maxRow = sheet.max_row

    totalCount = 0
    numIndex = 0
    for colIndex in range(1,maxCol+1):
        numCount = 0
        strCount = 0
        for rowIndex in range(1,maxRow+1):
            val = sheet.cell(row=rowIndex,column=colIndex).value

            if val == None:
                continue
            
            if isNum(val):
                numCount +=1
            else:
                if val.strip() != "":
                    strCount +=1
        if numCount >= 1 and strCount == 0:
            totalCount +=1
            numIndex = colIndex

    if totalCount == 0:
        raise Exception("需要有1列数字:"+sheetName)

    if totalCount >1:
        raise Exception("%s有%d列数字:" % (sheetName,totalCount))

    return numIndex
    
    

def loadBook(path):
    return openpyxl.reader.excel.load_workbook(path, read_only=True, \
        keep_vba=False, data_only=True, guess_types=False, keep_links=False)

def createBook():
    return openpyxl.workbook.workbook.Workbook(write_only=False)

def getVal(sheet,row,col):
    return sheet.cell(row=row,column=col).value

def setVal(sheet,row,col,val):
    sheet.cell(row=row,column=col).value = val

def getCell(sheet,row,col):
    return sheet.cell(row=row,column=col)

class MyCell(object):
    def __init__(self,val):
        self.cnt=1
        self.val=val

def getSameValues(sheet,col):
    ret = []
    sameCells = []
    for row in range(1,sheet.max_row):
        cell = getCell(sheet,row,col)
        if type(cell) == EmptyCell:
            continue
        f = False
        for i,sameCell in enumerate(sameCells):

            if sameCell.value == cell.value:
                rflag = False
                for idx,r in enumerate(ret):
                    if r.val == cell.value:
                        
                        r.cnt +=1
                        print(r.cnt)
                        sameCells[idx].cnt +=1
                        rflag = True
                        break
                if rflag == False:
                    same = MyCell(cell.value)
                    ret.append(same)
                f=True
                break
        if f == False:
            sameCells.append(cell)
    return ret
            

sheet1NumValSet = set()
sheet2NumValSet = set()
sheet1NumValList = []
sheet2NumValList = []
sheet1NumCellList = []
sheet2NumCellList = []

def do():
    global sheet1NumValSet
    global sheet2NumValSet
    global sheet1NumValList
    global sheet2NumValList
    try:
        if not os.path.isfile("in.xlsx"):
            raise Exception("in.xlsx不存在")
        if not os.access("in.xlsx",os.R_OK):
            raise Exception("in.xlsx不可读")


        inBook = loadBook("in.xlsx")
        
        if not isSheetNamesInBook(inBook,"Sheet1"):
            raise Exception("in.xlsx中不存在Sheet1")
        if not isSheetNamesInBook(inBook,"Sheet2"):
            raise Exception("in.xlsx中不存在Sheet2")

        inSheet1 = inBook["Sheet1"]
        
        inSheet2 = inBook["Sheet2"]
        
        sheet1NumColIndex = getNumColIndex(inSheet1)
        sheet2NumColIndex = getNumColIndex(inSheet2)

        outBook = createBook()
        outSheet = outBook.active
        outSheetCount = 0

        sheet1RowDone = {}
        sheet2RowDone = {}
        for sheet1RowIndex in range(1,inSheet1.max_row+1):
            sheet1RowDone[sheet1RowIndex]=False
            sheet1Cell = getCell(inSheet1,sheet1RowIndex,sheet1NumColIndex)
            a = None
            print(isinstance(outBook,Workbook))
        for sheet2RowIndex in range(1,inSheet2.max_row+1):
            sheet2RowDone[sheet2RowIndex]=False
            sheet2Cell = getCell(inSheet2,sheet2RowIndex,sheet2NumColIndex)
        
        

        thinSide = Side(border_style="thin", color="000000")
        thinBorder = Border(top=thinSide, left=thinSide, \
                            right=thinSide, bottom=thinSide)

        sheet1DiffRows = getDiffRows(inSheet1,sheet1NumColIndex)
        sheet2DiffRows = getDiffRows(inSheet2,sheet2NumColIndex)


        


        for sheet1RowIndex in sheet1DiffRows:
            if sheet1RowDone[sheet1RowIndex] == True:
                continue
            sheet1ValCell = getCell(inSheet1,sheet1RowIndex,\
                                        sheet1NumColIndex)
            if type(sheet1ValCell) == EmptyCell:
                continue
            for sheet2RowIndex in sheet2DiffRows:
                if sheet2RowDone[sheet2RowIndex] == True:
                    continue
                
                

                sheet2ValCell = getCell(inSheet2,sheet2RowIndex,\
                                        sheet2NumColIndex)
                if type(sheet2ValCell) == EmptyCell:
                    continue
                if sheet1ValCell.value == sheet2ValCell.value:
                    sheet1RowDone[sheet1RowIndex] = True
                    sheet2RowDone[sheet2RowIndex] = True
                    outSheetCount +=1
                    for sheet1ColIndex in range(1,inSheet1.max_column+1):
                        sheet1Cell = getCell(inSheet1,sheet1RowIndex,\
                                             sheet1ColIndex)
                        outCell = getCell(outSheet,outSheetCount,\
                                             sheet1ColIndex)
                        outCell.value = sheet1Cell.value

                        outCell.font = Font(color=colors.BLACK)
                        outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="CCFFFF",end_color="CCFFFF")
                        outCell.border=thinBorder;
                        

                    outSheetCount +=1
                    for sheet2ColIndex in range(1,inSheet2.max_column+1):
                        sheet2Cell = getCell(inSheet2,sheet2RowIndex,\
                                             sheet2ColIndex)
                        
                        outCell = getCell(outSheet,outSheetCount,\
                                             sheet2ColIndex)
                        outCell.value = sheet2Cell.value
                        
                        outCell.font = Font(color=colors.BLACK)
                        outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="FFCCFF",end_color="FFCCFF")
                        outCell.border=thinBorder;


        sheet1SameVals = getSameValues(inSheet1,sheet1NumColIndex)
        sheet2SameVals = getSameValues(inSheet2,sheet2NumColIndex)

        for sheet1SameVal in sheet1SameVals:
            print(sheet1SameVal.cnt)
        
        
        for sheet1RowIndex in range(1,inSheet1.max_row+1):
            if sheet1RowDone[sheet1RowIndex] == True:
                
                continue
            
            outSheetCount +=1
            for sheet1ColIndex in range(1,inSheet1.max_column+1):
                sheet1Cell = getCell(inSheet1,sheet1RowIndex,\
                                             sheet1ColIndex)
                outCell = getCell(outSheet,outSheetCount,\
                                             sheet1ColIndex)
                outCell.value = sheet1Cell.value

                outCell.font = Font(color=colors.BLACK)
                outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="CCFFFF",end_color="CCFFFF")
                outCell.border=thinBorder;
                        


        for sheet2RowIndex in range(1,inSheet2.max_row+1):
            if sheet2RowDone[sheet2RowIndex] == True:
                continue
                        

            outSheetCount +=1
            for sheet2ColIndex in range(1,inSheet2.max_column+1):
                sheet2Cell = getCell(inSheet2,sheet2RowIndex,\
                                             sheet2ColIndex)
                        
                outCell = getCell(outSheet,outSheetCount,\
                                             sheet2ColIndex)
                outCell.value = sheet2Cell.value
                        
                outCell.font = Font(color=colors.BLACK)
                outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="FFCCFF",end_color="FFCCFF")
                outCell.border=thinBorder;
        
        
        outBook.save("out.xlsx")
        inBook.close()
        outBook.close()
        print("success")
    except Exception:
        print(traceback.format_exc())

    time.sleep(2)
        
    
if __name__=="__main__":
    do()

