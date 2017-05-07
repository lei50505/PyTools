#! /usr/bin/env python
# -*- coding: utf-8 -*-

import sys
sys.path.append(".")

from openpyxl import *
from openpyxl.styles import *
from openpyxl.styles.numbers import *

import os


def toNum(val):
    if val is None:
        return None
        
    if isinstance(val,float):
        return val

    if isinstance(val,int):
        return float(val)

    if isinstance(val,str):
        try:
            return float(val)
        except:
            return None
    return None

class Cell():
    def __init__(self,cell):
        if cell is None:
            return
        self.cell = cell

    def getVal(self):
        if self.cell is None:
            return None
        return self.cell.value

    def getNumVal(self):
        val = self.getVal()
        if val is None:
            return None
        
        return toNum(val)

    def getStrVal(self):
        
        val = self.getVal()
        if val is None:
            return None
        try:
            strVal = str(val)
            if strVal.strip() != "":
                return strVal.strip()
        except:
            return None
        return None
    
    def setVal(self, val):
        if self.cell is None:
            return
        self.cell.value = val

    def setTextFormat(self):
        if self.cell is None:
            return 1
        self.cell.number_format = numbers.FORMAT_TEXT
        return 0
        

    def setRed(self):
        
        if self.cell is None:
            return 1
        self.cell.fill = PatternFill(fill_type = "solid",\
                            start_color="FFCCFF",end_color="FFCCFF")

        self.setBorder()

        return 0

    def setBlue(self):
        if self.cell is None:
            return 1
        self.cell.fill = PatternFill(fill_type = "solid", \
                            start_color="CCFFFF", end_color="CCFFFF")
        self.setBorder()

        return 0
        
    def setBorder(self):
        if self.cell is None:
            return 1
        thinSide = Side(border_style="thin", color="000000")
        thinBorder = Border(top=thinSide, left=thinSide, \
                            right=thinSide, bottom=thinSide)
        self.cell.border=thinBorder
        return 0
        

class Sheet():
    def __init__(self, sheet):
        self.sheet = sheet
        if self.sheet is None:
            return
        
        self.maxRow = sheet.max_row
        self.maxCol = sheet.max_column
        
        self.numColIndex = None
        
        self.numValDict = None
        self.numValSet = None
        self.numValList = None
        
        self.diffNumRows = None
        
        self.numRowDict = None
        self.numRowList = None

       


    def cell(self,row,col):
        if self.sheet is None:
            return None
        sheetCell = self.sheet.cell(row=row,column=col)
        cell = Cell(sheetCell)
        return cell

    def initNumColIndex(self):
        if self.sheet is None:
            return 1,"sheet为空"
        maxCol = self.maxCol
        maxRow = self.maxRow
        
        numColCount = 0
        numColIndex = 0
        
        for colIndex in range(1, maxCol + 1):
            numCellCount = 0
            isStrCell = False
            for rowIndex in range(1, maxRow + 1):
                cell = self.cell(rowIndex, colIndex)
                if cell is None:
                    continue
                numVal = cell.getNumVal()

                if isinstance(numVal,float):
                    numCellCount += 1
                    continue

                strVal = self.cell(rowIndex, colIndex).getStrVal()
                
                if isinstance(strVal,str):
                    isStrCell = True
                    break
                
            if numCellCount >= 1 and not isStrCell:
                numColCount += 1
                numColIndex = colIndex

        if numColCount == 1:
            self.numColIndex = numColIndex
            return 0,None

        if numColCount == 0:
            return 2,"没有数字列"

        if numColCount > 1:
            return 3,"有%d列是数字" % numColCount

        return 4,"初始化数字列失败"
        

    def initNumColDict(self):
        
        if self.numColIndex is None:
            return 1,"请先初始化numColIndex"
        
        self.numValDict = {}
        self.numValSet = set()
        self.numValList = []
        
        self.numRowDict = {}
        self.numRowList = []
        

        numColIndex = self.numColIndex
        maxRow = self.maxRow
        maxCol = self.maxCol
        
        for rowIndex in range(1, maxRow + 1):
            cell = self.cell(rowIndex, self.numColIndex)
            if cell is None:
                continue
            numVal = cell.getNumVal()
            if isinstance(numVal,float):
                self.numValSet.add(numVal)
                self.numValList.append(numVal)
                self.numRowDict[rowIndex] = numVal
                self.numRowList.append(rowIndex)
                
                dictVal = self.numValDict.get(numVal)
                if dictVal is None:
                    self.numValDict[numVal] = 1
                    continue
                
                self.numValDict[numVal] = dictVal + 1
        return 0,None
                
    def initDiffNumRows(self):
        if self.numRowList is None:
            return 1,"请先初始化numRowList"

        if self.numRowDict is None:
            return 2,"请先初始化numRowDict"

        if self.numValList is None:
            return 2,"请先初始化numValList"
        
        self.diffNumRows = []
        for numRow in self.numRowList:
            val = self.numRowDict.get(numRow)
            count = self.numValList.count(val)
            if count == 1:
                self.diffNumRows.append(numRow)
        return 0,None

    def getRowListByVal(self, val):
        if val is None:
            return None

        if self.numRowList is None:
            return None

        if self.numRowDict is None:
            return None
        
        val = toNum(val)
        if val is None:
            return None
        ret = []
        for numRow in self.numRowList:
            v = self.numRowDict.get(numRow)
            if v == val:
                ret.append(numRow)
        return ret

    def copyRowFromSheet(self, srcSheet, rowIndex):
        srcSheetMaxCol = srcSheet.maxCol
        
        for srcSheetColIndex in range(1, srcSheetMaxCol + 1):
            sheet2Cell = inSheet2.cell(sheet2RowIndex,\
                                             sheet2ColIndex)
            outCell = outSheet.cell(outSheetCount,\
                                             sheet2ColIndex)
            outCell.setVal(sheet2Cell.getVal())

            outCell.setRed()
            outCell.setTextFormat()
        
  

class Book():
    def __init__(self, book):
        self.book=book

    def active(self):
        if self.book is None:
            return 1,"book为空",None
        activeSheet = self.book.active
        return 0,None,Sheet(activeSheet)
        
    def sheet(self,sheetName):
        if self.book is None:
            return 1,"book没加载成功",None
        try:
            bookSheet = self.book[sheetName]
            sheet = Sheet(bookSheet)
            return 0,None,sheet
        except:
            pass
        return 2,"获取Sheet:%s失败" % sheetName,None

    def hasSheet(self, *sheetNames):

        if self.book is None:
            return False

        if len(sheetNames) == 0:
            return True
        
        bookSheetNames = self.book.get_sheet_names()
        for sheetName in sheetNames:
            if not isinstance(sheetName,str):
                return False
            if sheetName not in bookSheetNames:
                return False
        return True

    def save(self, path):
        if self.book is None:
            return 1
        if path is None:
            return 2
        self.book.save(path)

    def close(self):
        if self.book is not None:
            self.book.close()
        

def createBook():
    workBook =  Workbook(write_only=False)
    book = Book(workBook)
    return book

def loadBook(filePath):
    if not isinstance(filePath, str):
        return 1,"路径不是字符串类型",None
    if not os.path.isfile(filePath):
        return 2,"文件%s不存在" % filePath,None
    workBook = load_workbook(filePath, read_only = True, keep_vba = False, \
                    data_only = True, guess_types = False, keep_links = False)
    book = Book(workBook)
    return 0, None, book

if "__main__" == __name__:
    book = createBook()
    ret,err,sheet = book.active()
    print(sheet.maxCol)
    
