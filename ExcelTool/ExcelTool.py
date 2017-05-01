#! /usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import *
import openpyxl
import traceback
from openpyxl.styles import colors
from openpyxl.styles import Color,Font,Side,PatternFill,Border
from openpyxl.cell.read_only import EmptyCell,ReadOnlyCell
from openpyxl.workbook.workbook import Workbook
import time



def getRowsByVal(val,cellList):
    rows = []
    if not isinstance(val,float):
        return rows
    if not isinstance(cellList,list):
        return rows
    
    for i,cell in enumerate(cellList):
        v = getCellValIfNum(cell)
        if not isinstance(v,float):
            continue
        if v == val:
            rows.append(cell.row )
    return rows
    

def getDiffRowss(numSet,numList):
    if not isinstance(numSet,set):
        return None
    if not isinstance(numList,list):
        return None

    for i,num in enumerate(numList):
        pass

def getDiffRows(sheet,col):
    rowValid = {}
    diffCells = []
    ret = []
    for row in range(1,sheet.max_row+1):
        rowValid[row]=True
        
    for row in range(1,sheet.max_row+1):
        cell = getCell(sheet,row,col)
        if type(cell) == EmptyCell:
            continue
        findFlag = False
        equalRow = None
        for diffCell in diffCells:
            if diffCell.value == cell.value:
                findFlag= True
                equalRow = diffCell.row
                break

        if findFlag ==False:
            diffCells.append(cell)
        else:
            rowValid[row]=False
            rowValid[equalRow]=False
    for row in range(1,sheet.max_row+1):
        if rowValid[row] == True:
            ret.append(row)
    return ret
            
def getCellValIfNum(cell):
    
    if not isinstance(cell,ReadOnlyCell):
        return None

    val = cell.value

    return getValIfNum(val)



def isSheetNamesInBook(book,*sheetNames):
    if not isinstance(book,Workbook):
        return False
    
    bookSheetNames = book.get_sheet_names()
    for sheetName in sheetNames:
        if not isinstance(sheetName,str):
            return False
        if sheetName not in bookSheetNames:
            return False
    return True
        
    
def isNum(val):
    v = getValIfNum(val)
    if v is None:
        return False
    return True

def cellIsNum(cell):
    if cell == None:
        return False
    if type(cell) == EmptyCell:
        return False

    val = cell.value
    if val == None or str(val).strip() == "":
        return False
    try:
        int(val)
    except:
        try:
            float(val)
        except:
            return False
    return True

def getValIfNum(val):

    if isinstance(val,float):
        return val

    if isinstance(val,int):
        return float(val)

    if isinstance(val,str):
        try:
            return float(val)
            
        except:

            return None
    
    return None
    

def getNumColIndex(sheet):

    maxCol = sheet.max_column
    maxRow = sheet.max_row

    totalCount = 0
    numIndex = 0
    for colIndex in range(1,maxCol+1):
        numCount = 0
        strCount = 0
        for rowIndex in range(1,maxRow+1):
            val = sheet.cell(row=rowIndex,column=colIndex).value

            if val == None:
                continue
            
            if isNum(val):
                numCount +=1
            else:
                if val.strip() != "":
                    strCount +=1
        if numCount >= 1 and strCount == 0:
            totalCount +=1
            numIndex = colIndex

    if totalCount == 0:
        raise Exception("需要有1列数字:"+sheetName)

    if totalCount >1:
        raise Exception("%s有%d列数字:" % (sheetName,totalCount))

    return numIndex
    
    

def loadBook(path):
    return openpyxl.reader.excel.load_workbook(path, read_only=True, \
        keep_vba=False, data_only=True, guess_types=False, keep_links=False)

def createBook():
    return openpyxl.workbook.workbook.Workbook(write_only=False)

def getVal(sheet,row,col):
    return sheet.cell(row=row,column=col).value

def setVal(sheet,row,col,val):
    sheet.cell(row=row,column=col).value = val

def getCell(sheet,row,col):
    return sheet.cell(row=row,column=col)

class MyCell(object):
    def __init__(self,val):
        self.cnt=1
        self.val=val

def getSameValues(sheet,col):
    ret = []
    sameCells = []
    for row in range(1,sheet.max_row):
        cell = getCell(sheet,row,col)
        if type(cell) == EmptyCell:
            continue
        f = False
        for i,sameCell in enumerate(sameCells):

            if sameCell.value == cell.value:
                rflag = False
                for idx,r in enumerate(ret):
                    if r.val == cell.value:
                        
                        r.cnt +=1
                        print(r.cnt)
                        sameCells[idx].cnt +=1
                        rflag = True
                        break
                if rflag == False:
                    same = MyCell(cell.value)
                    ret.append(same)
                f=True
                break
        if f == False:
            sameCells.append(cell)
    return ret
            

sheet1NumValSet = set()
sheet2NumValSet = set()
sheet1NumValList = []
sheet2NumValList = []
sheet1NumCellList = []
sheet2NumCellList = []
sheet1MaxRow = 0
sheet1MaxCol = 0
sheet2MaxRow = 0
sheet2MaxCol = 0
inBook = None
outBook = None

def do():
    global sheet1NumValSet
    global sheet2NumValSet
    global sheet1NumValList
    global sheet2NumValList
    global sheet1MaxRow
    global sheet1MaxCol
    global sheet2MaxRow
    global sheet2MaxCol
    global inBook
    global outBook
    try:
        if not os.path.isfile("in.xlsx"):
            raise Exception("in.xlsx不存在")

        inBook = loadBook("in.xlsx")
        
        if not isSheetNamesInBook(inBook,"Sheet1"):
            raise Exception("in.xlsx中不存在Sheet1")
        if not isSheetNamesInBook(inBook,"Sheet2"):
            raise Exception("in.xlsx中不存在Sheet2")

        inSheet1 = inBook["Sheet1"]
        
        inSheet2 = inBook["Sheet2"]
        
        sheet1NumColIndex = getNumColIndex(inSheet1)
        sheet2NumColIndex = getNumColIndex(inSheet2)

        outBook = createBook()
        outSheet = outBook.active
        outSheetCount = 0

        sheet1RowDone = {}
        sheet2RowDone = {}

        sheet1MaxRow = inSheet1.max_row
        sheet1MaxCol = inSheet1.max_column
        sheet2MaxRow = inSheet2.max_row
        sheet2MaxCol = inSheet2.max_column
        
        for rowIndex in range(1,sheet1MaxRow+1):
            sheet1RowDone[rowIndex]=False
            cell = getCell(inSheet1,rowIndex,sheet1NumColIndex)
            val = getCellValIfNum(cell)
            if isinstance(val,float):
                sheet1NumValSet.add(val)
                sheet1NumValList.append(val)
                sheet1NumCellList.append(cell)
        for rowIndex in range(1,sheet2MaxRow+1):
            sheet2RowDone[rowIndex]=False
            cell = getCell(inSheet2,rowIndex,sheet2NumColIndex)
            val = getCellValIfNum(cell)
            if isinstance(val,float):
                sheet2NumValSet.add(val)
                sheet2NumValList.append(val)
                sheet2NumCellList.append(cell)

        print(sheet1NumValList)
        

        thinSide = Side(border_style="thin", color="000000")
        thinBorder = Border(top=thinSide, left=thinSide, \
                            right=thinSide, bottom=thinSide)

        sheet1DiffRows = getDiffRows(inSheet1,sheet1NumColIndex)
        sheet2DiffRows = getDiffRows(inSheet2,sheet2NumColIndex)


        


        for sheet1RowIndex in sheet1DiffRows:
            if sheet1RowDone[sheet1RowIndex] == True:
                continue
            sheet1ValCell = getCell(inSheet1,sheet1RowIndex,\
                                        sheet1NumColIndex)
            if type(sheet1ValCell) == EmptyCell:
                continue
            for sheet2RowIndex in sheet2DiffRows:
                if sheet2RowDone[sheet2RowIndex] == True:
                    continue
                
                

                sheet2ValCell = getCell(inSheet2,sheet2RowIndex,\
                                        sheet2NumColIndex)
                if type(sheet2ValCell) == EmptyCell:
                    continue
                if sheet1ValCell.value == sheet2ValCell.value:
                    sheet1RowDone[sheet1RowIndex] = True
                    sheet2RowDone[sheet2RowIndex] = True
                    outSheetCount +=1
                    for sheet1ColIndex in range(1,inSheet1.max_column+1):
                        sheet1Cell = getCell(inSheet1,sheet1RowIndex,\
                                             sheet1ColIndex)
                        outCell = getCell(outSheet,outSheetCount,\
                                             sheet1ColIndex)
                        outCell.value = sheet1Cell.value

                        outCell.font = Font(color=colors.BLACK)
                        outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="CCFFFF",end_color="CCFFFF")
                        outCell.border=thinBorder;
                        

                    outSheetCount +=1
                    for sheet2ColIndex in range(1,inSheet2.max_column+1):
                        sheet2Cell = getCell(inSheet2,sheet2RowIndex,\
                                             sheet2ColIndex)
                        
                        outCell = getCell(outSheet,outSheetCount,\
                                             sheet2ColIndex)
                        outCell.value = sheet2Cell.value
                        
                        outCell.font = Font(color=colors.BLACK)
                        outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="FFCCFF",end_color="FFCCFF")
                        outCell.border=thinBorder;


        sheet1SameVals = getSameValues(inSheet1,sheet1NumColIndex)
        sheet2SameVals = getSameValues(inSheet2,sheet2NumColIndex)

        for sheet1SameVal in sheet1SameVals:
            print(sheet1SameVal.cnt)
        
        
        for sheet1RowIndex in range(1,inSheet1.max_row+1):
            if sheet1RowDone[sheet1RowIndex] == True:
                
                continue
            
            outSheetCount +=1
            for sheet1ColIndex in range(1,inSheet1.max_column+1):
                sheet1Cell = getCell(inSheet1,sheet1RowIndex,\
                                             sheet1ColIndex)
                outCell = getCell(outSheet,outSheetCount,\
                                             sheet1ColIndex)
                outCell.value = sheet1Cell.value

                outCell.font = Font(color=colors.BLACK)
                outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="CCFFFF",end_color="CCFFFF")
                outCell.border=thinBorder;
                        


        for sheet2RowIndex in range(1,inSheet2.max_row+1):
            if sheet2RowDone[sheet2RowIndex] == True:
                continue
                        

            outSheetCount +=1
            for sheet2ColIndex in range(1,inSheet2.max_column+1):
                sheet2Cell = getCell(inSheet2,sheet2RowIndex,\
                                             sheet2ColIndex)
                        
                outCell = getCell(outSheet,outSheetCount,\
                                             sheet2ColIndex)
                outCell.value = sheet2Cell.value
                        
                outCell.font = Font(color=colors.BLACK)
                outCell.fill = PatternFill(fill_type = "solid",\
                            start_color="FFCCFF",end_color="FFCCFF")
                outCell.border=thinBorder;
        
        
        outBook.save("out.xlsx")
        inBook.close()
        outBook.close()
        print("success")
    except Exception:
        
        print(traceback.format_exc())
    finally:
        if isinstance(inBook,Workbook):
            inBook.close()
        if isinstance(outBook,Workbook):
            outBook.close()
        
        

    #time.sleep(2)
        
    
if __name__=="__main__":
    do()

