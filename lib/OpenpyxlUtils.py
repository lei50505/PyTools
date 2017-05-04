#! /usr/bin/env python
# -*- coding: utf-8 -*-

import sys
sys.path.append(".")

from openpyxl import *
from openpyxl.styles import *
from openpyxl.styles.numbers import *


class Cell():
    def __init__(self,cell):
        self.cell = cell
        self.val = self.cell.value

    def getNumVal(self):
        if isinstance(self.val,float):
            return self.val

        if isinstance(self.val,int):
            return float(self.val)

        if isinstance(self.val,str):
            try:
                return float(self.val)
            except:
                return None
            
        return None

    def getStrVal(self):
        if self.val is None:
            return None
        try:
            
            val = str(self.val)
            if val.strip()!="":
                return val.strip()
            
        except:
            return None
            
        return None
    def setVal(self,val):
        #self.cell.number_format = numbers.FORMAT_TEXT
        self.cell.value = val
        self.val = val
        

    def setRed(self):
        self.cell.fill = PatternFill(fill_type = "solid",\
                            start_color="FFCCFF",end_color="FFCCFF")

        self.setBorder()

    def setBlue(self):
        self.cell.fill = PatternFill(fill_type = "solid",\
                            start_color="CCFFFF",end_color="CCFFFF")
        self.setBorder()
        
    def setBorder(self):
        thinSide = Side(border_style="thin", color="000000")
        thinBorder = Border(top=thinSide, left=thinSide, \
                            right=thinSide, bottom=thinSide)
        self.cell.border=thinBorder
        

class Sheet():
    def __init__(self,sheet):
        self.sheet = sheet
        self.maxRow = sheet.max_row
        self.maxCol = sheet.max_column
        
        self.numColIndex = None
        self.__initNumColIndex__()

        self.numValDict = None
        self.numValSet = None
        self.numValList = None
        self.diffNumRows = None
        self.numRowDict = None
        self.numRowList = None
        self.__initNumColDict__()

    def cell(self,row,col):
        sheetCell = self.sheet.cell(row=row,column=col)
        cell = Cell(sheetCell)
        return cell

    def __initNumColIndex__(self):

        totalCount = 0
        numIndex = 0
        for colIndex in range(1,self.maxCol+1):
            numCount = 0
            hasStr = False
            for rowIndex in range(1,self.maxRow+1):
                numVal = self.cell(rowIndex,colIndex).getNumVal()

                if isinstance(numVal,float):
                    numCount +=1
                    continue

                strVal = self.cell(rowIndex,colIndex).getStrVal()
                
                if isinstance(strVal,str):
                    hasStr = True
                    break
            if numCount >= 1 and not hasStr:
                totalCount +=1
                numIndex = colIndex

        if totalCount == 0:
            return

        if totalCount >1:
            return

        self.numColIndex = numIndex

    def __initNumColDict__(self):
        if self.numColIndex is None:
            return
        self.numValDict = {}
        self.numValSet = set()
        self.numValList = []
        self.diffNumRows = []
        self.numRowDict = {}
        self.numRowList = []
        for rowIndex in range(1,self.maxRow+1):
            cell = self.cell(rowIndex,self.numColIndex)
            val = cell.getNumVal()
            if isinstance(val,float):
                self.numValSet.add(val)
                self.numValList.append(val)
                self.numRowDict[rowIndex] = val
                self.numRowList.append(rowIndex)
                dictVal = self.numValDict.get(val)
                if dictVal is None:
                    self.numValDict[val]=1
                    continue
                self.numValDict[val] = dictVal+1

        for numRow in self.numRowList:
            val = self.numRowDict.get(numRow)
            count = self.numValList.count(val)
            if count == 1:
                self.diffNumRows.append(numRow)


    def getRowListByVal(self,val):
        val = float(val)
        ret = []
        for numRow in self.numRowList:
            v = self.numRowDict.get(numRow)
            if v == val:
                ret.append(numRow)
        return ret
        
        

class Book():
    def __init__(self,book,mode):
        self.book=book

        self.active = None
        if mode == "create":
            activeSheet = book.active
            self.active = Sheet(activeSheet)

        
        
    def sheet(self,sheetName):
        bookSheet = self.book[sheetName]
        sheet = Sheet(bookSheet)
        return sheet

    def hasSheet(self,*sheetNames):
        
        bookSheetNames = self.book.get_sheet_names()
        for sheetName in sheetNames:
            if not isinstance(sheetName,str):
                return False
            if sheetName not in bookSheetNames:
                return False
        return True

    def save(self,path):
        self.book.save(path)

    def close(self):
        if isinstance(self.book,Workbook):
            self.book.close()
        

def createBook():
    workBook =  Workbook(write_only=False)
    book = Book(workBook,"create")
    return book

def loadBook(filePath):
    if not isinstance(filePath,str):
        return None
    workBook = load_workbook(filePath, read_only=True, keep_vba=False,\
                         data_only=True,guess_types=False, keep_links=False)
    book = Book(workBook,"load")
    return book



if __name__ == "__main__":
    book1 = createBook()
    book2 = loadBook("Test.xlsx")
    print(book2.hasSheet("Sheet1"))

    print(book2.sheet("Sheet1").numValDict)
    print(book2.sheet("Sheet1").numValSet)
    print(book2.sheet("Sheet1").numValList)
    print(book2.sheet("Sheet1").diffNumRows)
    print(book2.sheet("Sheet1").numRowDict)
    print(book2.sheet("Sheet1").numRowList)

    print(book2.sheet("Sheet1").getRowListByVal(54))

    sheet = book1.active
    sheet.cell(1,1).setBlue()

    book1.save("out.xlsx")

    book1.close()
    book2.close()

